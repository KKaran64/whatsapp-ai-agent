#!/usr/bin/env python3
import sys
try:
    import openpyxl
    has_openpyxl = True
except ImportError:
    has_openpyxl = False

try:
    import pandas as pd
    has_pandas = True
except ImportError:
    has_pandas = False

if not has_openpyxl and not has_pandas:
    print("ERROR: Neither openpyxl nor pandas is installed")
    print("Install with: pip3 install openpyxl")
    sys.exit(1)

filename = sys.argv[1]

if has_pandas:
    try:
        df = pd.read_excel(filename, sheet_name=None)
        for sheet_name, sheet_data in df.items():
            print(f"\n{'='*60}")
            print(f"SHEET: {sheet_name}")
            print(f"{'='*60}\n")
            print(sheet_data.to_string(index=False))
    except Exception as e:
        print(f"ERROR: {e}")
elif has_openpyxl:
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n{'='*60}")
            print(f"SHEET: {sheet_name}")
            print(f"{'='*60}\n")
            for row in sheet.iter_rows(values_only=True):
                print('\t'.join([str(cell) if cell is not None else '' for cell in row]))
    except Exception as e:
        print(f"ERROR: {e}")
